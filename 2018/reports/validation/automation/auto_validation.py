import pandas as pd
import openpyxl as xl
import openpyxl.utils.dataframe as xldf
import datetime, os, re, locale, sys
import locale
locale.setlocale(locale.LC_ALL,'')

# model_dir = "E:\Projects\Clients\NashvilleMPO\ModelUpdate2023\Model\Development\nashabm_TCAD9_transit"

# HIGHWAY VALIDATION
def highway_validation(model_dir, sce_dir):
    # Read template
    # print(model_dir)
    hwy_tmpl = xl.reader.excel.load_workbook(
        filename=os.path.join(model_dir,
                              'Support/BaseYear/validation/hwy_validation_template.xlsx'),
        )
    wksht = hwy_tmpl['assignment_result']

    # Read model outputs
    results = pd.read_csv(
        os.path.join(
            model_dir, sce_dir,
            'outputs/assignment_result_3.csv')
            ).set_index('ID').sort_index()

    # ensure all IDs are present in new data and data is ordered
    # for row in wksht.iter_rows(min_row=2):
    #     assert row[2].value in results.index.values and row[2].value == results.index[row[2].row - 2], f"Missing ID: {row[2].value} at row {row[2].row - 2}"

    # Convert model outputs to Excel sheet of same size as template
    results_rows = xldf.dataframe_to_rows(results.reset_index(),index=False)
    nwb = xl.Workbook()
    nws = nwb.active
    for row in results_rows:
        nws.append(row)
    nws.insert_cols(1,2)
    columns = {col[0].column : col[0].value for col in nws.iter_cols(min_row=1,max_row=1,min_col=4)}

    # Overwrite data from the template w/ new data
    for row in nws.iter_rows():
        for column in columns.keys():
            wksht[row[column-1].coordinate].value = row[column-1].value

    # Output new workbook
    validation_dir = os.path.join(model_dir, sce_dir, 'reports/validation/')
    # Create the directory if it doesn't exist
    os.makedirs(validation_dir, exist_ok=True)

    hwy_tmpl.save(
        os.path.join(
            validation_dir,
            f'hwy_validation_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx'))

# TRANSIT VALIDATION
def transit_validation(model_dir, sce_dir):
    # Read template
    trn_tmpl = xl.reader.excel.load_workbook(
        filename=os.path.join(model_dir,'Support/BaseYear/validation/trn_validation_template.xlsx'),
        )
    wksht = trn_tmpl['TrnStat.asc']

    # Read model outputs and overwrite template data
    with open(os.path.join(model_dir, sce_dir, 'outputs/TrnStat.asc')) as results:
        row_counter = 1
        
        for line in results:
            col_counter = 0
            for val in re.split(' +',line):
                try:
                    # Convert string to appropriate float value
                    wksht[row_counter][col_counter].value = locale.atof(val)
                except ValueError:
                    # If it's a string, sanitize it for Excel
                    wksht[row_counter][col_counter].value = re.sub("=","",val)
                col_counter += 1
            row_counter += 1

    # Output new workbook

    # Output new workbook
    validation_dir = os.path.join(model_dir, sce_dir, 'reports/validation/')
    # Create the directory if it doesn't exist
    os.makedirs(validation_dir, exist_ok=True)

    trn_tmpl.save(
        os.path.join(
            validation_dir,
            f'trn_validation_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx'))
    
if __name__=="__main__":
    if len(sys.argv) < 3:
        print("Usage: python validation_summaries.py <model_dir>")
        sys.exit(-1)
        
    model_dir = sys.argv[1]
    sce_dir = sys.argv[2]
    
    highway_validation(model_dir, sce_dir)
    transit_validation(model_dir, sce_dir)

    print('done')
